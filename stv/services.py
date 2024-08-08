from django.contrib.auth.models import User

from django.conf import settings
from datetime import datetime
from openpyxl import load_workbook


def create_new_report(instance, request):
    direccion = request.POST.get('direccion')
    administrador = request.POST.get('administrador')
    otro = request.POST.get('otros')
    certificado_norma = request.POST.get('certificado_norma')
    contrato = request.POST.get('contrato')
    equipo = request.POST.get('equipo')
    fecha_mantenimiento = request.POST.get('fecha_mantenimiento')

    fecha_hoy = datetime.now().strftime('%Y-%m-%d')
    file_path = f'{settings.STATICFILES_DIRS[0]}/stv_files/formato_visita.xlsx'
    output_path = f'{settings.MEDIA_ROOT}/reportes/{instance.cliente.title()} - {fecha_hoy}.xlsx'

    wb = load_workbook(file_path)
    replace_name_reporte_xlsx(wb, file_path, f'input_{instance.tipo_reporte}', 'X' if not otro else otro)
    replace_name_reporte_xlsx(wb, file_path, 'input_cliente', instance.cliente.title())
    replace_name_reporte_xlsx(wb, file_path, 'input_ciudad', instance.ciudad)
    replace_name_reporte_xlsx(wb, file_path, 'input_direccion', direccion)
    replace_name_reporte_xlsx(wb, file_path, 'input_administrador', administrador)
    replace_name_reporte_xlsx(wb, file_path, 'input_certificado_norma', certificado_norma)
    replace_name_reporte_xlsx(wb, file_path, 'input_contrato', contrato)
    replace_name_reporte_xlsx(wb, file_path, 'input_equipo', equipo)
    replace_name_reporte_xlsx(wb, file_path, 'input_mantenimiento', fecha_mantenimiento)

    clean_empty_inputs(wb)
    wb.save(output_path)

    instance.archivo.name = f'{instance.cliente.title()} - {fecha_hoy}.xlsx'
    instance.save()
    return instance


def replace_name_reporte_xlsx(wb, file_path, target_word, replacement_word):
    for sheet in wb.sheetnames:
        ws = wb[sheet]

        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and target_word in cell.value:
                    cell.value = cell.value.replace(target_word, replacement_word)
    return wb


def clean_empty_inputs(wb):
    for sheet in wb.sheetnames:
        ws = wb[sheet]

        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and 'input' in cell.value:
                    cell.value = ''
    return wb


def create_new_empleado_user(nombre, apellido):
    username = f"{nombre.split(' ')[0].lower()}.{apellido.split(' ')[0].lower()}"
    user = User.objects.create(username=username)
    user.set_password('Omega2030*')
    user.save()
    return user
